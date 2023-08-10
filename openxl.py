!pip install pydata_google_auth
!pip install webdriver_manager
!pip install selenium
!pip install openpyxl
!pip install xlsxwriter

import pandas as pd
import gspread
import pydata_google_auth
import time
import os
import xlsxwriter 
import openpyxl
from openpyxl.styles import PatternFill 
from datetime import date


scope = ['https://www.googleapis.com/auth/spreadsheets']

# json_file_name = json_file_name
credentials = pydata_google_auth.get_user_credentials(scope, auth_local_webserver=True)

#구글스프레드시트 불러오기 
gs_c = gspread.authorize(credentials)
spreadsheet_url = 'https://docs.google.com/spreadsheets/d/1svWlC99jP3pb93ROVBBfYCBNwgtk7KUszuUETNRrXuw/edit#gid=0'
doc = gs_c.open_by_url(spreadsheet_url)
s1 = doc.worksheet('해지요청_포켓파이')
s2 = doc.worksheet('해지요청_마트너')
s3 = doc.worksheet('해지요청_CNRS100')

df_1 = pd.DataFrame(s1.get_all_records())
df_2 = pd.DataFrame(s2.get_all_records())
df_3 = pd.DataFrame(s3.get_all_records())

# 복사 범위 지정
ex_col_1 = df_1.iloc[:, 0:7]
ex_col_2 = df_2.iloc[:, 0:7]
ex_col_3 = df_3.iloc[:, 0:7]

# 오늘 날짜를 문자열로 변환
today = date.today().strftime('%Y%m%d')

# 파일 이름에 오늘 날짜를 추가하여 저장
file1 = f'[SOCAR] SKT 회선 해지 요청_포켓파이_{today}.xlsx'
file2 = f'[SOCAR] SKT 회선 해지 요청_마트너_{today}.xlsx'
file3 = f'[SOCAR] SKT 회선 해지 요청_CNR S100C_{today}.xlsx'

# 작업 디렉토리 변경
os.chdir('C:/Users/austin/Downloads')

# 엑셀 파일로 저장
ex_col_1.to_excel(file1, index=False)
ex_col_2.to_excel(file2, index=False)
ex_col_3.to_excel(file3, index=False)

# 엑셀 불러오기 
wb1 = openpyxl.load_workbook(file1)
wb2 = openpyxl.load_workbook(file2)
wb3 = openpyxl.load_workbook(file3)

#시트 선택 
ws1 = wb1['Sheet1']
ws2= wb2['Sheet1']
ws3= wb3['Sheet1']


#열 너비 보정
col_fix = ws1.column_dimensions["B"].width = 15
col_fix = ws1.column_dimensions["F"].width = 15
col_fix = ws1.column_dimensions["G"].width = 15
col_fix = ws1.column_dimensions["H"].width = 15

col_fix = ws2.column_dimensions["B"].width = 15
col_fix = ws2.column_dimensions["F"].width = 15
col_fix = ws2.column_dimensions["E"].width = 15
col_fix = ws2.column_dimensions["G"].width = 15
col_fix = ws2.column_dimensions["H"].width = 15

col_fix = ws3.column_dimensions["B"].width = 15
col_fix = ws3.column_dimensions["F"].width = 15
col_fix = ws3.column_dimensions["E"].width = 15
col_fix = ws3.column_dimensions["G"].width = 15
col_fix = ws3.column_dimensions["H"].width = 15


#개통번호 보정
last_row = len(ex_col_1)
i = 2 

for i in range(2, last_row + 2) : 
    ws1.cell(row = i, column = 8, value = str('0' + str(ws1.cell(row = i, column = 2).value)))
    
    i+=1

last_row = len(ex_col_2)
i = 2 

for i in range(2, last_row + 2) : 
    ws2.cell(row = i, column = 8, value = str('0' + str(ws2.cell(row = i, column = 2).value)))
    
    i+=1
    
last_row = len(ex_col_3)
i = 2 

for i in range(2, last_row + 2) : 
    ws3.cell(row = i, column = 8, value = str('0' + str(ws3.cell(row = i, column = 2).value)))
    
    i+=1

# 4) 엑셀 저장하기
wb1.save(file1)
wb2.save(file2)
wb3.save(file3)